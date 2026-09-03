"""
batch_xlsx_reader.py — чтение входного .xlsx для batch-режима (07-batch-runner).

Проверяет наличие обязательного столбца «Название компании» и желательных
«Ссылка на hh.ru», «Количество открытых вакансий», «Ссылка на сайт»
(п.3.1.2 ТЗ). Отсутствие необязательных столбцов — предупреждение, а не
ошибка: 07-batch-runner должен уметь работать с минимальным входом и сам
доопределять сайт/hh.ru через 01-company-resolve.

Использование:
    python batch_xlsx_reader.py --file companies.xlsx
"""

from __future__ import annotations

import argparse

import openpyxl

from common import emit, envelope, fail

REQUIRED_COLUMNS = {"название компании"}
OPTIONAL_COLUMNS = {
    "ссылка на hh.ru": "hh_url",
    "количество открытых вакансий": "open_vacancies_hint",
    "ссылка на сайт": "site",
}
REQUIRED_KEY = "company"


def normalize_header(cell_value: object) -> str:
    return str(cell_value or "").strip().lower()


def main() -> None:
    parser = argparse.ArgumentParser(description="read and validate batch input .xlsx")
    parser.add_argument("--file", required=True)
    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file, data_only=True)
    except FileNotFoundError:
        fail(f"файл не найден: {args.file}", code=3)
        return
    except Exception as e:  # noqa: BLE001 — битый xlsx — ошибка аргументов, а не источника
        fail(f"не удалось открыть {args.file} как .xlsx: {e}", code=3)
        return

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        emit(
            envelope(
                ok=False,
                source="batch_xlsx_reader",
                method="openpyxl",
                url=args.file,
                error={"type": "config_error", "message": "файл пуст — нет строки заголовков"},
            )
        )
        return

    headers = [normalize_header(h) for h in header_row]
    col_index = {h: i for i, h in enumerate(headers) if h}

    missing_required = REQUIRED_COLUMNS - set(col_index.keys())
    if missing_required:
        emit(
            envelope(
                ok=False,
                source="batch_xlsx_reader",
                method="openpyxl",
                url=args.file,
                error={
                    "type": "config_error",
                    "message": f"отсутствует обязательный столбец: {', '.join(missing_required)}",
                },
            )
        )
        return

    warnings = []
    present_optional = {}
    for col_name, key in OPTIONAL_COLUMNS.items():
        if col_name in col_index:
            present_optional[key] = col_index[col_name]
        else:
            warnings.append(f"необязательный столбец «{col_name}» отсутствует — будет определяться автоматически")

    company_idx = col_index["название компании"]
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if row is None or all(v is None for v in row):
            continue
        company = str(row[company_idx]).strip() if company_idx < len(row) and row[company_idx] is not None else ""
        if not company:
            warnings.append(f"строка {idx + 1}: пустое название компании — пропущена")
            continue
        entry = {"idx": idx, "company": company}
        for key, col_i in present_optional.items():
            value = row[col_i] if col_i < len(row) else None
            entry[key] = str(value).strip() if value is not None else None
        rows.append(entry)

    if not rows:
        warnings.append("после фильтрации не осталось ни одной валидной строки")

    emit(
        envelope(
            ok=len(rows) > 0,
            source="batch_xlsx_reader",
            method="openpyxl",
            url=args.file,
            data={"rows": rows, "columns_found": sorted(col_index.keys())},
            warnings=warnings,
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # noqa: BLE001
        emit(envelope(ok=False, source="batch_xlsx_reader", method="openpyxl", error={"type": "unexpected_error", "message": str(e)}))

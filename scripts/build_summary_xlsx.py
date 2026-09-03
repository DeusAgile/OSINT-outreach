"""
build_summary_xlsx.py — сводная выгрузка по итогам batch-прогона (п.4.1 ТЗ).

Читает run-state.json (чекпоинт 07-batch-runner) и строит .xlsx с двумя
листами:
- Summary: гиперссылки на markdown-отчёты по каждой компании + ключевые
  поля прямо в таблице (сфера деятельности, найденные ФИО и контакты,
  confidence), freeze panes на первой строке, autofilter.
- Находки: построчно все найденные персоны по всем компаниям, с заливкой
  ячейки confidence по уровню (CONFIRMED/FIRM/TENTATIVE).

Использование:
    python build_summary_xlsx.py --state run-state.json --out summary.xlsx

Читает опциональные поля напрямую из run-state.json построчно: report_path,
persons_found, max_confidence, industry, contacts, status, error,
needs_review (bool, строка или {chosen, alternate, reason} — см.
modules/01-company-resolve.md п.4d — компания определена автоматически из
нескольких равнозначных кандидатов, требует проверки человеком; подсвечена
отдельной колонкой, а не блокирует прогон вопросом). Модуль 01/05/06 обязаны
писать их в run-state.json по ходу прогона — этот скрипт только оформляет
то, что там уже есть, и не обращается к markdown-отчётам напрямую (чтобы не
дублировать парсинг между модулем и скриптом).
"""

from __future__ import annotations

import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from common import emit, envelope, fail

CONFIDENCE_FILL = {
    "CONFIRMED": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "FIRM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "TENTATIVE": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

NEEDS_REVIEW_FILL = PatternFill(start_color="FFD9A0", end_color="FFD9A0", fill_type="solid")

SUMMARY_HEADERS = [
    "Компания", "Статус", "Отчёт", "Сфера деятельности", "Найдено персон",
    "Макс. confidence", "Контакты", "Требует проверки", "Ошибка",
]
FINDINGS_HEADERS = ["Компания", "ФИО", "Должность", "Confidence", "Источник факта"]


def autosize(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)


def build(state: dict, out_path: str) -> dict:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(SUMMARY_HEADERS)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    summary_ws.freeze_panes = "A2"

    findings_ws = wb.create_sheet("Находки")
    findings_ws.append(FINDINGS_HEADERS)
    for cell in findings_ws[1]:
        cell.font = Font(bold=True)
    findings_ws.freeze_panes = "A2"

    widths_summary = {i: len(h) for i, h in enumerate(SUMMARY_HEADERS, start=1)}
    widths_findings = {i: len(h) for i, h in enumerate(FINDINGS_HEADERS, start=1)}

    warnings = []
    rows = state.get("rows", [])
    for row in rows:
        company = row.get("company", "")
        status = row.get("status", "pending")
        report_path = row.get("report_path")
        industry = row.get("industry", "")
        persons_found = row.get("persons_found", 0)
        max_confidence = row.get("max_confidence", "")
        contacts = ", ".join(row.get("contacts", []) or [])
        error = row.get("error", "")
        needs_review = row.get("needs_review")
        if isinstance(needs_review, dict):
            chosen = needs_review.get("chosen", "")
            alternate = needs_review.get("alternate", "")
            reason = needs_review.get("reason", "")
            review_text = f"выбрано: {chosen}; альтернатива: {alternate}" + (f" ({reason})" if reason else "")
        elif needs_review:
            review_text = str(needs_review)
        else:
            review_text = ""

        r = [company, status, None, industry, persons_found, max_confidence, contacts, review_text, error]
        summary_ws.append(r)
        row_idx = summary_ws.max_row
        if report_path:
            cell = summary_ws.cell(row=row_idx, column=3, value=report_path)
            cell.hyperlink = report_path
            cell.font = Font(color="0563C1", underline="single")
        else:
            summary_ws.cell(row=row_idx, column=3, value="—")
        if review_text:
            summary_ws.cell(row=row_idx, column=8).fill = NEEDS_REVIEW_FILL

        for i, val in enumerate(r, start=1):
            widths_summary[i] = max(widths_summary.get(i, 0), len(str(val or "")))

        for person in row.get("persons", []) or []:
            frow = [
                company,
                person.get("name", ""),
                person.get("position", ""),
                person.get("confidence", ""),
                person.get("evidence_url", ""),
            ]
            findings_ws.append(frow)
            frow_idx = findings_ws.max_row
            fill = CONFIDENCE_FILL.get(person.get("confidence", ""))
            if fill:
                findings_ws.cell(row=frow_idx, column=4).fill = fill
            for i, val in enumerate(frow, start=1):
                widths_findings[i] = max(widths_findings.get(i, 0), len(str(val or "")))

    if not rows:
        warnings.append("run-state.json не содержит строк (rows) — сводная таблица пуста")

    autosize(summary_ws, widths_summary)
    autosize(findings_ws, widths_findings)
    summary_ws.auto_filter.ref = summary_ws.dimensions
    findings_ws.auto_filter.ref = findings_ws.dimensions

    wb.save(out_path)
    return {"out": out_path, "rows_written": len(rows), "warnings": warnings}


def main() -> None:
    parser = argparse.ArgumentParser(description="build summary.xlsx from run-state.json")
    parser.add_argument("--state", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    try:
        with open(args.state, "r", encoding="utf-8") as f:
            state = json.load(f)
    except FileNotFoundError:
        fail(f"файл состояния не найден: {args.state}", code=3)
        return
    except json.JSONDecodeError as e:
        fail(f"{args.state} содержит невалидный JSON: {e}", code=3)
        return

    result = build(state, args.out)
    emit(
        envelope(
            ok=True,
            source="build_summary_xlsx",
            method="openpyxl",
            url=args.out,
            data={"rows_written": result["rows_written"]},
            warnings=result["warnings"],
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # noqa: BLE001
        emit(envelope(ok=False, source="build_summary_xlsx", method="openpyxl", error={"type": "unexpected_error", "message": str(e)}))
